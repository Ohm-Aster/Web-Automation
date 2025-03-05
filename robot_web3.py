import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# Cargar archivo Excel y seleccionar la hoja de trabajo
EXCEL_FILE = "./datos.xlsx"
SHEET_NAME = "Exitoso"

wb = load_workbook(EXCEL_FILE)
matric = wb[SHEET_NAME]
wb.close()

# Credenciales y URL
USERS = ['MiUsuario']
PASSWORD = 'MiConstraseña'
URL = 'https://encuestasregresoseguro.com/'

# Inicializar navegador
driver = webdriver.Chrome()
driver.maximize_window()
driver.get(URL)

def iniciar_sesion(user):
    """Inicia sesión en la plataforma con el usuario dado."""
    driver.find_element(By.XPATH, '//*[@id="tipoEscuela_id"]/option[3]').click()
    driver.find_element(By.XPATH, '//*[@id="universidades"]/option[15]').click()
    driver.find_element(By.CSS_SELECTOR, '#DNI').send_keys(user)
    driver.find_element(By.CSS_SELECTOR, '#formAcceso input[type="password"]').send_keys(PASSWORD)
    driver.find_element(By.CSS_SELECTOR, '#botonEnviar').click()

def validar_matriculas():
    """Valida las matrículas cargadas desde el archivo Excel."""
    boton_enviar = '#formbusqueda button.btn-success'
    boton_confirmar = '.swal2-confirm'
    boton_reintentar = 'section.content-header center a'

    exito = 0

    for row in matric.iter_rows(min_row=2, max_row=1474, min_col=2, max_col=3, values_only=True):
        nombre, matricula = row
        if not matricula:
            continue

        campo_matricula = driver.find_element(By.CSS_SELECTOR, '#matricula')
        campo_matricula.clear()
        campo_matricula.send_keys(matricula, Keys.RETURN)

        time.sleep(1)  # Espera para evitar errores de carga
        driver.find_element(By.CSS_SELECTOR, boton_enviar).click()
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR, boton_confirmar).click()

        if driver.find_elements(By.CSS_SELECTOR, boton_reintentar):
            exito += 1
            driver.find_element(By.CSS_SELECTOR, boton_reintentar).click()

    print(f"Validaciones exitosas: {exito}")
    driver.find_element(By.CSS_SELECTOR, "aside .info a i").click()

# Ejecutar el proceso
iniciar_sesion(USERS[1])  # Usa el segundo usuario
validar_matriculas()

driver.quit()
